# Description: FM.py contains the FileManager class. Contents:
from pathlib import Path
from datetime import datetime, timedelta
import json
import csv
import re
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from termcolor import colored
from Logicbomb import Logic

class FileManager:
    @staticmethod
    def load_creds(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                creds = yaml.safe_load(f)
            return creds
        except FileNotFoundError:
            print(colored(f"Error: {filename} file not found.", "red", attrs=["bold"]))
            return None
        except yaml.YAMLError as e:
            print(colored(f"Error: Failed to load {filename} file. {e}", "red", attrs=["bold"]))
            return None

    @staticmethod
    def generate_filename(file_path):
        """ Generate a unique filename by incrementing the version number until an available name is found.
        :param file_path: The base Path object for the file, including extension but excluding version.
        :return: A Path object for the file with a unique version number if the file already exists."""
        version = 1
        base_name = file_path.stem
        extension = file_path.suffix
        directory = file_path.parent
        base_name = re.sub(r'_v\d+$', '', base_name)  # Remove any existing version number in the filename
        while True:
            new_file_path = directory / f"{base_name}_v{version}{extension}"
            if not new_file_path.exists():
                return new_file_path
            version += 1

    @staticmethod
    def check_json_content(filepath):
        json_file = Path(filepath)
        if json_file.is_file() and json_file.stat().st_size > 0:
            try:
                with json_file.open('r', encoding='utf-8') as f:
                    data = json.load(f)
                    return bool(data)
            except json.JSONDecodeError:
                return False  # JSON file is corrupted
        return False

    @staticmethod
    def get_date_properties():
        current_date = datetime.now()
        current_date_str = current_date.strftime("%d-%m-%Y")
        ninety_days_ago = current_date - timedelta(days=90)
        ninety_days_ago_str = ninety_days_ago.strftime("%d-%m-%Y")
        print(colored(f"Today's date: {current_date_str} | Date 90 days ago string: {ninety_days_ago_str}", "yellow", attrs=["bold"]))
        print(colored(f"Following dateTime will be used for logic bomb: {ninety_days_ago}", "red", attrs=["bold"]))
        return current_date_str, ninety_days_ago_str, ninety_days_ago

    @staticmethod
    def json_to_csv(filepath_json, filepath_csv, filepath_csv_filtered, filepath_csv_disabled, logger):
        with open(filepath_json, 'r', encoding='utf-8-sig') as fp:
            raw_data = json.load(fp)
        df = pd.json_normalize(raw_data, sep='_')
        df = df.drop(['id', 'onPremisesSecurityIdentifier', 'securityIdentifier','signInActivity_lastNonInteractiveSignInRequestId','signInActivity_lastSignInRequestId'], axis=1)
        datetime_columns = [
            'signInSessionsValidFromDateTime',
            'createdDateTime',
            'onPremisesLastSyncDateTime',
            'refreshTokensValidFromDateTime',
            'signInActivity_lastSignInDateTime',
            'signInActivity_lastNonInteractiveSignInDateTime']
        for col in datetime_columns:  # Remove time from datetime columns
            if col in df.columns:
                df[col] = df[col].astype(str).str.split('T', expand=True)[0]
        custom_order = [
            'userType',
            'displayName',
            'mail',
            'signInActivity_lastSignInDateTime',
            'signInActivity_lastNonInteractiveSignInDateTime',
            'signInSessionsValidFromDateTime',
            'refreshTokensValidFromDateTime',
            'onPremisesLastSyncDateTime',
            'createdDateTime',
            'userPrincipalName',
            'otherMails',
            'onPremisesUserPrincipalName',
            'passwordProfile_forceChangePasswordNextSignIn',
            'passwordProfile_forceChangePasswordNextSignInWithMfa',
            'passwordPolicies',
            'externalUserState',
            'externalUserStateChangeDateTime',
            'onPremisesSamAccountName',
            'onPremisesSyncEnabled',
            'onPremisesDistinguishedName',
            'identities',
            'passwordProfile',
            'passwordProfile_password',
            'usageLocation',
            'onPremisesDomainName',
            'accountEnabled',]
        missing_columns = [col for col in custom_order if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Columns {missing_columns} are missing in DataFrame after dropping columns.")
        df = df[custom_order]  # Apply the custom order
        if 'accountEnabled' in df.columns:
            enabled_df = df[df['accountEnabled']].copy()
            disabled_df = df[~df['accountEnabled']].copy()
        else:
            raise ValueError("Column 'accountEnabled' not found in DataFrame.")
        _, ninety_days_ago_str, ninety_days_ago = FileManager.get_date_properties()
        logger.info(f"Any user with no sign-in date before {ninety_days_ago_str} is non-compliant.")
        # Use the get_row_identifier function when calling check_compliance
        compliant_df = enabled_df[enabled_df.apply(
            lambda row: Logic.check_compliance(row, datetime_columns, ninety_days_ago, logger, Logic.get_row_identifier(row)),
            axis=1
        )]
        non_compliant_df = enabled_df[~enabled_df.index.isin(compliant_df.index)]
        compliant_df.to_csv(filepath_csv, index=False, quoting=csv.QUOTE_ALL)
        non_compliant_df.to_csv(filepath_csv_filtered, index=False, quoting=csv.QUOTE_ALL)
        disabled_df.to_csv(filepath_csv_disabled, index=False, quoting=csv.QUOTE_ALL)
        logger.info(f"Total Count: {len(df)}")
        logger.info(f"Enabled Account Count: {len(enabled_df)} | Disabled Account Count: {len(disabled_df)}")
        logger.info(f"Compliant User Count: {len(compliant_df)} | Non-Compliant User Count: {len(non_compliant_df)}")

    @staticmethod
    def excel_fmt_header(workbook, latest_files, logger, excel_writer):
        fmt_header = workbook.add_format({'bold': True, 'bg_color': 'blue', 'color': 'white', 'font_name': 'Open Sans', 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1})
        for base_name, (csv_file, _) in latest_files.items():
            df = pd.read_csv(csv_file, dayfirst=True)
            if df.empty:
                logger.info(f"The DataFrame for {csv_file.name} is empty.")
                continue
            df.to_excel(excel_writer, sheet_name=base_name, index=False)
            worksheet = excel_writer.sheets[base_name]
            for col_num, value in enumerate(df.columns.values):  # Apply header format
                worksheet.write(0, col_num, value, fmt_header)

    @staticmethod
    def csv_to_excel(output_dir, excel_filename, logger):
        excel_filepath = Path(output_dir) / excel_filename
        latest_files = FileManager.find_csv_files_in_dir(output_dir, logger)
        with pd.ExcelWriter(excel_filepath, engine='xlsxwriter') as excel_writer:  # pylint: disable=abstract-class-instantiated
            workbook = excel_writer.book
            FileManager.excel_fmt_header(workbook, latest_files, logger, excel_writer)

    @staticmethod
    def find_csv_files_in_dir(output_dir, logger):
        latest_files = {}
        for csv_file in Path(output_dir).glob('*.csv'):
            match = re.match(r"(.+?)_v(\d+)\.csv", csv_file.name)
            if match:
                base_name, version = match.groups()
                version = int(version)
                if base_name not in latest_files or version > latest_files[base_name][1]:
                    latest_files[base_name] = (csv_file, version)
            else:
                logger.info(f"File {csv_file.name} does not match the expected pattern.")
        for base_name, (csv_file, version) in latest_files.items():  # Log the latest files before returning
            logger.info(f"Latest file for {base_name}: {csv_file.name} with version {version}")
        if not latest_files:
            logger.info("No CSV files were found matching the expected versioned pattern.")
        return latest_files

    @staticmethod
    def modify_excel_file(excel_filename, logger):
        wb = load_workbook(excel_filename)
        # Font definitions
        common_font = Font(name='Open Sans', size=10, color="c8d1d9")
        bad_cell_fmt = Font(name='Open Sans', size=10, color="FF0000")
        good_cell_fmt = Font(name='Open Sans', size=10, color="00FF00")
        purple_font = Font(name='Open Sans', size=10, color="800080")
        yellow_font = Font(name='Open Sans', size=10, color="FFD700")
        # Alignment definition
        alignment = Alignment(horizontal='center', vertical='center')
        # Fill definitions - updated for better contrast
        dark_green_fill = PatternFill(start_color="004d00", end_color="004d00", fill_type="solid")
        dark_yellow_fill = PatternFill(start_color="806000", end_color="806000", fill_type="solid")
        light_pink_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        default_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        light_lavender_fill = PatternFill(start_color="F4E6FF", end_color="F4E6FF", fill_type="solid")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2):  # Apply common font and alignment, skip header
                for cell in row:
                    cell.font = common_font
                    cell.fill = default_fill  # Apply default fill to all cells
                    cell.alignment = alignment
            for cell in sheet['A'][1:]:  # Apply font colors based on column A values, skip header
                if cell.value == "Member":
                    cell.font = purple_font
                    cell.fill = light_lavender_fill
                elif cell.value == "Guest":
                    cell.font = yellow_font
                    cell.fill = dark_yellow_fill
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:  # Set the font and background color for columns D-I based on dates, skip header
                for cell in sheet[col][1:]:
                    if cell.value is not None:
                        if isinstance(cell.value, datetime):
                            cell_date = cell.value
                        elif isinstance(cell.value, str):
                            try:
                                cell_date = datetime.strptime(cell.value, '%Y-%m-%d')
                            except ValueError:
                                logger.error(f"Date format error in cell {cell.coordinate}: {cell.value}")
                                continue
                        else:
                            logger.error(f"Unknown cell value type in cell {cell.coordinate}: {cell.value}")
                            continue
                        if (datetime.now() - cell_date).days <= 90:
                            cell.font = good_cell_fmt
                            cell.fill = dark_green_fill
                        else:
                            cell.font = bad_cell_fmt
                            cell.fill = light_pink_fill
                    else:
                        cell.value = "NA"
                        cell.font = common_font

        wb._sheets.append(wb._sheets.pop(1)) # type: ignore  # Move sheet 2 to the last spot

        # Save the workbook
        try:
            wb.save(excel_filename)
            logger.info("Excel file formatted and saved successfully.")
        except Exception as e:
            logger.error("An error occurred while saving the Excel file: " + str(e))
